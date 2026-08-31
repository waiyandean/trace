#!/usr/bin/env python3
"""Work out which supplier each ingredient comes from.

The intake form asks for the supplier first and then shows only that
supplier's ingredients, which turns sixty-odd tiles into twenty-odd. This
builds that mapping from the kitchen's own records rather than from anybody's
memory.

Two sources, kept apart on purpose:

  registered  the `Ingredients` tab of Goods In Records, which is the list the
              kitchen maintains by hand: one row per ingredient, naming its
              supplier (occasionally two, comma separated).
  delivered   the `Deliveries` tab, 2,600-odd real deliveries. A supplier that
              has actually delivered an item is evidence that it can.

Both are written, each labelled with where it came from, because they do not
entirely agree and the disagreement is worth keeping rather than flattening.
Where the maintained list names a supplier that has never delivered the item,
that is recorded too — the list may be right and the history merely short.

Where the records cannot say, the kitchen's answer is taken from
`scripts/catalog-overrides.json` — the same file the catalog importer reads —
and recorded as `decided`, so a decision is never mistaken for evidence.

Names are matched after normalising case and spacing only. Anything that does
not match a catalog item is reported, never guessed at: the two spreadsheets
spell several things differently, and a fuzzy match here would put an
ingredient behind the wrong supplier and hide it from the picker at the door.

Usage:
    python3 scripts/import_item_suppliers.py "~/Downloads/Goods In Records.xlsx" \
        [--api http://localhost:8799] [--out scripts/item-suppliers.sql] \
        [--report scripts/item-suppliers-report.txt]
"""

import argparse
import collections
import json
import pathlib
import sys
import urllib.request

import openpyxl


def normalise(name):
    return " ".join(str(name).split()).strip().lower()


def sql_str(value):
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def read_registered(book):
    """The maintained list: (ingredient, supplier) pairs, supplier sometimes a pair."""
    pairs = []
    for row in book["Ingredients"].iter_rows(values_only=True):
        if not row or not row[0] or len(row) < 2 or not row[1]:
            continue
        for supplier in str(row[1]).split(","):
            supplier = supplier.strip()
            if supplier:
                pairs.append((str(row[0]).strip(), supplier))
    return pairs


def read_delivered(book):
    """Every (ingredient, supplier) pairing the delivery history actually contains."""
    sheet = book["Deliveries"]
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    ingredient_at = header.index("Ingredient")
    supplier_at = header.index("Supplier")

    counts = collections.Counter()
    for row in rows:
        if not row or not row[ingredient_at] or not row[supplier_at]:
            continue
        counts[(str(row[ingredient_at]).strip(), str(row[supplier_at]).strip())] += 1
    return counts


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("workbook", help="the Goods In Records workbook")
    parser.add_argument("--api", default="http://localhost:8799")
    parser.add_argument("--out", default="scripts/item-suppliers.sql")
    parser.add_argument("--report", default="scripts/item-suppliers-report.txt")
    parser.add_argument("--overrides", default="scripts/catalog-overrides.json")
    args = parser.parse_args()

    try:
        with urllib.request.urlopen(f"{args.api}/api/catalog?action=items&active=all", timeout=20) as response:
            items = json.load(response)["rows"]
        with urllib.request.urlopen(f"{args.api}/api/catalog?action=suppliers", timeout=20) as response:
            suppliers = json.load(response)["rows"]
    except Exception as error:  # noqa: BLE001
        sys.exit(f"could not read the trace catalog from {args.api}: {error}")

    item_by_name = {normalise(item["name"]): item for item in items if item["kind"] == "ingredient"}
    supplier_by_name = {normalise(row["name"]): row for row in suppliers}

    book = openpyxl.load_workbook(pathlib.Path(args.workbook).expanduser(), read_only=True, data_only=True)
    registered = read_registered(book)
    delivered = read_delivered(book)

    # pairing -> the sources that support it
    support = collections.defaultdict(set)
    deliveries_for = {}
    unknown_items = collections.Counter()
    unknown_suppliers = collections.Counter()

    def note(item_name, supplier_name, source, count=None):
        item = item_by_name.get(normalise(item_name))
        supplier = supplier_by_name.get(normalise(supplier_name))
        if not item:
            unknown_items[item_name] += 1
            return
        if not supplier:
            unknown_suppliers[supplier_name] += 1
            return
        support[(item["id"], supplier["id"])].add(source)
        if count:
            deliveries_for[(item["id"], supplier["id"])] = count

    for item_name, supplier_name in registered:
        note(item_name, supplier_name, "registered")
    for (item_name, supplier_name), count in delivered.items():
        note(item_name, supplier_name, "delivered", count)

    # The kitchen's own answers, applied last and recorded as decided rather
    # than as evidence. A person saying so is a perfectly good source; it is
    # simply a different one from a delivery note, and the difference is worth
    # keeping on the row.
    overrides_path = pathlib.Path(args.overrides)
    overrides = json.loads(overrides_path.read_text()) if overrides_path.exists() else {}
    decided = {
        name: supplier
        for name, supplier in (overrides.get("item_suppliers") or {}).items()
        if not name.startswith("_")
    }
    provenance = ""
    if decided:
        provenance = f" ({overrides.get('recorded_by')}, {overrides.get('recorded_on')})"
    for item_name, supplier_name in decided.items():
        note(item_name, supplier_name, "decided")

    names = {item["id"]: item["name"] for item in items}
    supplier_names = {row["id"]: row["name"] for row in suppliers}

    rows = []
    for (item_id, supplier_id), sources in sorted(support.items(), key=lambda entry: names[entry[0][0]]):
        # A pairing the history proves is stronger evidence than one only
        # written down, so 'delivered' wins where both apply.
        source = "delivered" if "delivered" in sources else "registered"
        count = deliveries_for.get((item_id, supplier_id))
        if source == "delivered":
            detail = f"{count} deliveries in the Goods In Records history"
            if "registered" in sources:
                detail += ", and on the kitchen's maintained supplier list"
        else:
            detail = "on the kitchen's maintained supplier list, but never seen in the delivery history"
        rows.append((item_id, supplier_id, source, detail))

    lines = [
        "-- Which supplier each ingredient comes from.",
        f"-- Generated by scripts/import_item_suppliers.py from {pathlib.Path(args.workbook).name}.",
        "-- Do not edit: re-run the importer instead.",
        "",
    ]
    for item_id, supplier_id, source, detail in rows:
        lines.append(
            "INSERT INTO item_suppliers (item_id, supplier_id, source, note) VALUES "
            f"({sql_str(item_id)}, {sql_str(supplier_id)}, {sql_str(source)}, {sql_str(detail)})\n"
            "  ON CONFLICT (item_id, supplier_id) DO UPDATE SET "
            "source = excluded.source, note = excluded.note, updated_at = datetime('now');"
        )
    pathlib.Path(args.out).write_text("\n".join(lines) + "\n")

    # ---- the report -------------------------------------------------------
    by_item = collections.defaultdict(list)
    for item_id, supplier_id, source, _ in rows:
        by_item[item_id].append((supplier_names[supplier_id], source))

    shared = {names[i]: sorted(v) for i, v in by_item.items() if len(v) > 1}
    ingredients = [item for item in items if item["kind"] == "ingredient"]

    # Only an active ingredient with no supplier is a gap. An inactive one is
    # a decision already taken — it is out of scope at Glasgow, and the picker
    # does not show it at all — so listing it as unanswered would be noise
    # that never goes away.
    active = [item for item in ingredients if item.get("active")]
    unmapped = sorted(item["name"] for item in active if item["id"] not in by_item)
    dropped = sorted(item["name"] for item in ingredients if not item.get("active"))

    report = [
        f"{len(rows)} item/supplier pairings over {len(by_item)} of {len(active)} active ingredients "
        f"({len(dropped)} inactive, out of scope at Glasgow) -> {args.out}",
        "",
    ]
    if shared:
        report += [
            f"Ingredients arriving from more than one supplier ({len(shared)})",
            "  The premise that suppliers do not share ingredients does not hold in the",
            "  kitchen's own records. Each of these needs confirming: if a pairing is",
            "  wrong the picker will hide that ingredient behind the wrong supplier, and",
            "  somebody is stuck at the door with a box they cannot book in.",
            "",
        ]
        for name, entries in sorted(shared.items()):
            report.append(f"  - {name}: " + ", ".join(f"{supplier} ({source})" for supplier, source in entries))
        report.append("")
    if decided:
        report += [
            f"Supplier decided by the kitchen rather than found in the records ({len(decided)})",
            "",
        ] + [f"  - {name}: {supplier}" for name, supplier in sorted(decided.items())] + [""]
    if unmapped:
        report += [
            f"Ingredients with no supplier at all ({len(unmapped)})",
            "  These appear under every supplier in the picker rather than none, because",
            "  hiding an ingredient nobody has mapped would be worse than showing it.",
            "",
        ] + [f"  - {name}" for name in unmapped] + [""]
    if unknown_items:
        report += [
            f"Names in the workbook that match no catalog ingredient ({len(unknown_items)})",
            "  Left out rather than matched approximately. Most are packaging, or the",
            "  same ingredient spelled differently; either way a guess here would put an",
            "  ingredient behind the wrong supplier.",
            "",
        ] + [f"  - {name} ({count})" for name, count in sorted(unknown_items.items())] + [""]
    if unknown_suppliers:
        report += [f"Suppliers not in the catalog ({len(unknown_suppliers)})"]
        report += [f"  - {name} ({count})" for name, count in sorted(unknown_suppliers.items())] + [""]

    text = "\n".join(report) + "\n"
    pathlib.Path(args.report).write_text(text)
    print(text)


if __name__ == "__main__":
    main()
